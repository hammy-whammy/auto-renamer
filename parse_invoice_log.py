#!/usr/bin/env python3
"""Parse PDF renaming log file and create Excel report with original filename, new filename, and status"""

import re
import pandas as pd
from pathlib import Path
import sys
from datetime import datetime

def parse_log_file(log_path):
    """Parse the log file and extract renaming information"""
    
    results = []
    
    with open(log_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Split by processing entries - look for "Processing file X/Y: filename.pdf"
    processing_pattern = r'Processing file \d+/\d+: (.+?\.pdf)'
    processing_matches = re.finditer(processing_pattern, content)
    
    for match in processing_matches:
        original_filename = match.group(1)
        start_pos = match.start()
        
        # Find the next processing entry or end of file
        next_match = None
        for next_processing in re.finditer(processing_pattern, content[start_pos + 1:]):
            next_match = next_processing
            break
        
        if next_match:
            end_pos = start_pos + 1 + next_match.start()
            entry_content = content[start_pos:end_pos]
        else:
            entry_content = content[start_pos:]
        
        # Initialize entry data
        entry = {
            'Original Filename': original_filename,
            'New Filename': '',
            'Status': 'Unknown'
        }
        
        # Check for success patterns (both dry-run and live modes)
        success_pattern_dry = r'✅ SUCCESS: Would rename \'([^\']+)\' → \'([^\']+)\''
        success_pattern_live = r'✅ SUCCESS: Renamed \'([^\']+)\' → \'([^\']+)\''
        
        success_match = re.search(success_pattern_dry, entry_content) or re.search(success_pattern_live, entry_content)
        
        if success_match:
            entry['Status'] = 'SUCCESSFUL'
            entry['New Filename'] = success_match.group(2)
        
        # Check for failure pattern
        elif '❌ FAILED:' in entry_content:
            entry['Status'] = 'FAILED'
            entry['New Filename'] = 'N/A'
        
        # Check for skipped pattern
        elif '⏭️ SKIPPED:' in entry_content:
            entry['Status'] = 'SKIPPED'
            entry['New Filename'] = 'N/A'
        
        results.append(entry)
    
    return results

def create_excel_report(results, output_path):
    """Create Excel report from parsed results"""
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Ensure we have the exact columns requested
    df = df[['Original Filename', 'New Filename', 'Status']]
    
    # Create Excel writer with styling
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='PDF Renaming Results', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['PDF Renaming Results']
        
        # Add some basic formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Status-based coloring
        success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        failed_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')   # Light red
        skipped_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Light yellow
        
        # Apply status coloring and center alignment
        status_col = 3  # Status column is the 3rd column (C)
        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
            status_cell = worksheet.cell(row=row, column=status_col)
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if status_cell.value == 'SUCCESSFUL':
                for col in range(1, 4):  # Columns A, B, C
                    worksheet.cell(row=row, column=col).fill = success_fill
            elif status_cell.value == 'FAILED':
                for col in range(1, 4):  # Columns A, B, C
                    worksheet.cell(row=row, column=col).fill = failed_fill
            elif status_cell.value in ['SKIPPED', 'Unknown']:
                for col in range(1, 4):  # Columns A, B, C
                    worksheet.cell(row=row, column=col).fill = skipped_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters for long filenames
            worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function"""
    
    # Get log file path from command line or use default
    if len(sys.argv) > 1:
        log_path = Path(sys.argv[1])
    else:
        # Use the most recent log file
        log_dir = Path('logs')
        if not log_dir.exists():
            print("❌ Logs directory not found")
            return
        
        log_files = list(log_dir.glob('pdf_renaming_*.log'))
        if not log_files:
            print("❌ No log files found")
            return
        
        # Get the most recent log file
        log_path = max(log_files, key=lambda f: f.stat().st_mtime)
    
    if not log_path.exists():
        print(f"❌ Log file not found: {log_path}")
        return
    
    print(f"📄 Parsing log file: {log_path}")
    
    # Parse the log file
    try:
        results = parse_log_file(log_path)
        print(f"📊 Found {len(results)} processing entries")
        
        # Create output filename based on log file name
        log_timestamp = log_path.stem.replace('pdf_renaming_', '')
        output_path = f"invoice_tracking_report_{log_timestamp}.xlsx"
        
        # Create Excel report
        create_excel_report(results, output_path)
        
        print(f"✅ Excel report created: {output_path}")
        
        # Print summary
        successful_count = sum(1 for r in results if r['Status'] == 'SUCCESSFUL')
        failed_count = sum(1 for r in results if r['Status'] == 'FAILED')
        skipped_count = sum(1 for r in results if r['Status'] in ['SKIPPED', 'Unknown'])
        
        print(f"\n📈 SUMMARY:")
        print(f"   ✅ Successful: {successful_count}")
        print(f"   ❌ Failed: {failed_count}")
        print(f"   ⏭️  Skipped/Unknown: {skipped_count}")
        print(f"   📋 Total: {len(results)}")
        
        # Show first few examples
        if results:
            print(f"\n📋 FIRST 3 EXAMPLES:")
            for i, result in enumerate(results[:3], 1):
                status_icon = "✅" if result['Status'] == 'SUCCESSFUL' else "❌" if result['Status'] == 'FAILED' else "⏭️"
                print(f"   {i}. {status_icon} {result['Original Filename']} → {result['New Filename']} ({result['Status']})")
        
    except Exception as e:
        print(f"❌ Error processing log file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
